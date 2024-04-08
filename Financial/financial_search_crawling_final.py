from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
import urllib.request
import urllib
import time
import math
import os
import random
from openpyxl import load_workbook
from openpyxl.styles import Alignment

while True:   
    #서두
    print("="*50)
    print("\t     증권 뉴스 검색기 입니다!")
    print("="*50)
    print("\n")
    
    query_txt = input("원하는 검색어를 입력하세요 : ")

    #크롬 드라이버 옵션 설정
    options = webdriver.ChromeOptions()
    options.add_argument("headless")
    options.add_argument("disable-gpu")
    driver = webdriver.Chrome(options=options)
    
    #네이버 증권 검색
    driver.get('https://finance.naver.com/news/news_search.naver')
    s_time = time.time()
    time.sleep(1)
    
    element_N = driver.find_element(By.CLASS_NAME, "inputTxt")
    driver.find_element(By.CLASS_NAME, "inputTxt").click()
    element_N.send_keys(query_txt)
    element_N.send_keys("\n")
    time.sleep(1)
    
    now_url = driver.current_url
    
    result_cnt = driver.find_element(By.XPATH,'//*[@id="contentarea_left"]/div[2]/p/strong[2]').text
    
    if result_cnt == "0":
            print("검색 결과가 없습니다.\n 다시 검색 해주세요")
    
    else:
        print("="*50)
        print("\t   총 검색 결과는 "+result_cnt+"개 입니다.")
        print("="*50)
        print("\n")
        
        want_no = input("몇 개의 기사를 저장할까요? ")
        articles = round(int(want_no)/20)
        
        #1-1. 네이버 내용 추출
        def clean_txt(element):
            return element.get_text().strip().replace('\n', '').replace('\t', '')
        
        cnt_N = 1
        count_N = [] #기사갯수
        date_N = [] #날짜
        time_N = [] #시간
        title_N = [] #기사 제목
        head_N = [] #헤드라인
        from_N = [] #출처
        link_N = [] #뉴스링크
        
        for i in range(1, articles+2):
            html1_N = driver.page_source
            soup1_N = BeautifulSoup(html1_N, 'html.parser')
            content1_N = soup1_N.find_all(class_='articleSubject')
            content2_N = soup1_N.find_all('dd','articleSummary')
        
            for a in content1_N :
                count_N.append(cnt_N)
                cnt_N += 1
            
                title2_N = clean_txt(a)
                title_N.append(title2_N)
            
                link2_N = a.find('a')['href']
                link_N.append(link2_N)
        
            for b in content2_N :
                date2_N = clean_txt(b.find(class_="wdate"))
                date_N.append(date2_N[:10])
                time_N.append(date2_N[10:])
            
                head2_N = clean_txt(b)
                head_N.append(head2_N)
        
                from2_N = clean_txt(b.find(class_="press"))
                from_N.append(from2_N)
        
            if i!=0:
                page_url = now_url+"&page="+str(i)
                driver.get(page_url)
        
            time.sleep(0.5)
        
        #엑셀 파일 만들고 추가하기
        import pandas as pd
        from datetime import date
        
        today1 = date.today()
        today = today1.strftime("%Y-%m-%d")
        naver_list = pd.DataFrame()
        
        naver_list['번호'] = count_N[:int(want_no)]
        naver_list['날짜'] = date_N[:int(want_no)]
        naver_list['시간'] = time_N[:int(want_no)]
        naver_list['제목'] = title_N[:int(want_no)]
        naver_list['내용'] = head_N[:int(want_no)]
        naver_list['출처'] = from_N[:int(want_no)]
        naver_list['사이트'] = link_N[:int(want_no)]
        
        fx_name = "c:\\personal_python\\Financial\\seaching_"+query_txt+"_"+today+".xlsx"
        naver_list.to_excel(fx_name , index = False, engine='openpyxl')
        
        #엑셀 파일 디자인
        wb = load_workbook("c:\\personal_python\\Financial\\seaching_"+query_txt+"_"+today+".xlsx", data_only=True)
        ws = wb['Sheet1']
        
        ws.column_dimensions['A'].width = 5.0
        ws.column_dimensions['B'].width = 15.5
        ws.column_dimensions['C'].width = 6.0
        ws.column_dimensions['D'].width = 55.0
        ws.column_dimensions['E'].width = 65.0
        ws.column_dimensions['F'].width = 10.0
        ws.column_dimensions['G'].width = 50.0
        
        for column in ['A', 'B', 'C']:
            for cell in ws[column]:
                cell.alignment = Alignment(horizontal='center')
        
        wb.save("c:\\personal_python\\Financial\\seaching_"+query_txt+"_"+today+".xlsx")
        
        #데이터 시각화
        from wordcloud import WordCloud
        import matplotlib.pyplot as plt
        import re
        import konlpy
        from konlpy.tag import Okt
        from collections import Counter

        title_word = []

        #배열에 제목을 저장 후 특문 제거
        for cell in ws['D']:
            cell.value = re.sub(r'[^\nㄱ-힣\s]', '', cell.value)
            title_word.append(cell.value)
        
        #명사만 잘라서 저장 후 한 글자 배열 삭제
        okt = Okt()
        nouns = []
        for i in title_word:
            nouns.append(okt.nouns(i))

        def remove_single_char_strings(lst):
            return [s for s in lst if len(s) != 1]
        remove_single_char_strings(nouns)

        #명사들을 txt파일에 저장
        nouns_list = ' '.join([elem for sublist in nouns for elem in sublist])
        with open(query_txt+'.txt', 'w', encoding='utf-8') as f:
            f.write(nouns_list)

        #워드클라우드로 객체 비율 보기
        text = open(query_txt+'.txt', 'r', encoding='utf-8').read()
        wordcloud = WordCloud(max_words=100,font_path = "C:\personal_python\Financial\MaruBuri-SemiBold.ttf",
                              background_color='white',colormap="Accent_r",width=800,height=800).generate(text)

        #시각화
        plt.figure(figsize=(10,10))
        plt.imshow(wordcloud)
        plt.axis('off')
        plt.savefig(query_txt+'.png')

        e_time = time.time()
        t_time = e_time - s_time
        print("총 소요시간은 %s 초 입니다" %round(t_time,1))
        wb.close()
        driver.quit()
        break